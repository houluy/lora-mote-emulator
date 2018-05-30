import openpyxl
import sys


class Test_data:
    def __init__(self, filename):
        self.filename = filename
        try:
            wb = openpyxl.load_workbook(filename=filename)
        except Exception as e:
            print(e)
            sys.exit(0)
        self.sheet = wb[wb.sheetnames[0]]

    def _get_numeric_column(self, column):
        return ord(column) - ord('A')

    @staticmethod
    def form_pos(column, row):
        return '{column}{row}'.format(column=column, row=row)

    def range_values(self, start_col='A', end_col='A', start_row=1, end_row=1):
        start_col = self._get_numeric_column(start_col)
        end_col = self._get_numeric_column(end_col)
        cells = self.sheet.iter_rows(
            max_col=end_col + 1,
            min_col=start_col,
            max_row=end_row,
            min_row=start_row
        )
        row_number = end_row - start_row + 1
        # col_number = end_col - start_col + 1
        values = [[] for x in range(row_number)]
        for row, c in enumerate(cells):
            for cell in c:
                if cell.value:
                    values[row].append(cell.value)
                else:
                    values[row].append('')
        return {
            'title': values[0],
            'value': values[1:],
        }
